from django.db.models import Sum
from .models import *
import os.path
from openpyxl import Workbook, load_workbook
from datetime import datetime
from django.utils.timezone import localdate

def init():
    if now().date() <= Updated.objects.all().last().date:
        return
      
    '''
    Player.objects.all().delete()
    Stats.objects.all().delete()
    Team.objects.all().delete()
    Member.objects.all().delete()
    Game.objects.all().delete()
    '''
    
    i = 0
    while True:
        i += 1
        file_path = 'matches/'+str(i)+'.xlsx'
        if os.path.exists(file_path) is False:
            i-=1      
            read_futures(i)
            u = Updated(date = now().date(), last_excel_read = i)
            u.save()
            break
        
        book = load_workbook(file_path)
        sheet = book.active
        
        host = sheet['C1'].value
        guest = sheet['D1'].value
        
        if i % 2 == 0:
            current_team_name = guest
        else:
            current_team_name = host
            
        try:
            current_team = Team.objects.get(name = current_team_name)
        except Team.DoesNotExist:
            newTeam = Team(name = current_team_name)
            newTeam.save()
            current_team = newTeam 
            
        date = datetime.date(datetime.strptime( sheet['B1'].value, "%Y-%m-%d"))
        
        if i % 2 == 0:
            try:
                # match = Game.objects.filter(date=date).get(guest=current_team)
                    
                match = Game.objects.filter(date=date).get(guest=current_team)
                if match.played == True:
                    continue #should I return here???
                else:
                    match.played = True
                    match.save()
                    continue
            except Game.DoesNotExist:
                newMatch = Game(date = date)
                newMatch.host = Team.objects.get(name = host)
                newMatch.guest = Team.objects.get(name = guest)
                newMatch.played = True
                newMatch.save()
        
        index = 2
        while sheet['B'+str(index)].value is not None:
            s = Stats()
            
            s.team = current_team_name
            s.date = date
            
            player_name  = sheet['B'+str(index)].value
            try:
                s.player = Player.objects.get(name = player_name)
            except Player.DoesNotExist:
                newPlayer = Player(name = player_name)
                newPlayer.save()
                s.player = newPlayer
                
            try:
                member = Member.objects.filter(player=s.player).get(team=current_team)
                if member.date_start > s.date:
                    member.date_start = s.date
                    member.save()
            except Member.DoesNotExist:
                newMember = Member(player=s.player, team=current_team, date_start = s.date)
                newMember.save()
            
            s.Pos   = sheet['C'+str(index)].value
            s.Rank  = sheet['D'+str(index)].value
            s.PPI   = sheet['E'+str(index)].value
            s.MP    = sheet['F'+str(index)].value
            s.G     = sheet['G'+str(index)].value
            s.SOnT  = sheet['H'+str(index)].value 
            s.SOffT = sheet['I'+str(index)].value
            s.BS    = sheet['J'+str(index)].value
            s.OG    = sheet['K'+str(index)].value
            s.A     = sheet['L'+str(index)].value
            s.P     = sheet['M'+str(index)].value
            s.C     = sheet['N'+str(index)].value
            s.Tk    = sheet['O'+str(index)].value
            s.INT   = sheet['P'+str(index)].value
            s.FW    = sheet['Q'+str(index)].value
            s.FC    = sheet['R'+str(index)].value
            s.O     = sheet['S'+str(index)].value
            s.YC    = sheet['T'+str(index)].value
            s.RC    = sheet['U'+str(index)].value
            s.GC    = sheet['V'+str(index)].value
            s.PW    = sheet['W'+str(index)].value
            s.SAV   = sheet['X'+str(index)].value
            s.Ca    = sheet['Y'+str(index)].value
            s.KS    = sheet['Z'+str(index)].value
            
            try:
                exist = Stats.objects.filter(player=s.player).get(date=s.date)
            except Stats.DoesNotExist:
                s.save()
            index += 1
    
    
        
def read_futures(i=0):
    #if now().date() <= Updated.objects.all().last().date:
    #    return 

    file_path = 'future/future-matches.xlsx'
    if os.path.exists(file_path) is False:
        return
        
    book = load_workbook(file_path)
    sheet = book.active
    
    sheet['A1'].value = i

    index = 1
    while sheet['B'+str(index)].value is not None:
        date  = sheet['B'+str(index)].value
        
        
        host = sheet['C'+str(index)].value
        guest = sheet['D'+str(index)].value
            
        
            
        date = datetime.date(datetime.strptime(sheet['B'+str(index)].value, "%Y-%m-%d"))
        
        
        try:
            match = Game.objects.filter(date=date).get(guest__name=guest)
        except Game.DoesNotExist:
            newMatch = Game(date = date)
            newMatch.host = Team.objects.get(name = host)
            newMatch.guest = Team.objects.get(name = guest)
            newMatch.played = False
            newMatch.save()

        index += 1